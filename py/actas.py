from openpyxl import load_workbook
from campusvirtual import CampusVirtual
from actascalificacion import ActasCalificacion
import argparse, os, requests, re
from selenium import webdriver

grade_ceil = False

def calificar_actas(notas, actas, prefijo, requisito, scale):
    wb = load_workbook(filename=notas, read_only=True)
    wba = [load_workbook(filename=fname) for fname in actas]
    clean_workbooks(wba)
    for wbi in wba:
        calificacion_por_defecto(wbi)
    required_columns = calcular_columnas(wb.active, requisito)
    for row in wb.active:
        givenName, sn, nota = (cell.value for cell in (row[0], row[1], row[6]))
        if givenName == 'Nombre':
            continue
        no_presentado = any(row[i].value == '-' for i in required_columns)
        nota = 0. if nota == '-' else calificacion(nota, scale)
        if no_presentado and nota < 5.:
            continue
        for wbi in wba:
            poner_nota(wbi, '{}, {}'.format(sn,givenName), nota)

    for wbi, fname in zip(wba, actas):
        wbi.save(prefijo + fname)

def clean_workbooks(wbs):
    ''' Elimina estilos empotrados que incluyen las actas.
        Si no se hace openpyxl genera un archivo incorrecto.
    '''
    for wb in wbs:
        for row in wb.active:
            for cell in row:
                cell.style = 'Normal'

def calcular_columnas(sheet, requisito):
    header = [cell.value for cell in next(iter(sheet))]
    return [posicion_actividad(header, r) for r in requisito]

def posicion_actividad(L, e):
    ra = re.compile(r'\w+:(?P<nombre>[\w\s]+)\s\(\w*\)')
    for i,t in enumerate(L):
        match = ra.match(t)
        if not match:
            continue
        if match.group('nombre') == e:
            return i


def calificacion_por_defecto(wb):
    for row in wb.active:
        if row[0].value.startswith('Número'):
            continue
        row[2].value = None
        row[3].value = 'NP'

def poner_nota(wb, nombre, nota):
    for row in wb.active:
        if row[1].value.upper() == nombre:
            row[2].value = round(nota,1)
            row[2].number_format = '0.0'
            row[3].value = letra(nota)
            break

def calificacion(nota, scale):
    v = float(nota)/scale
    if grade_ceil and v < 5. and v > 4.:
        return 5.
    return v

def letra(v):
    if v < 5.:
        return 'SS'
    if v < 7.:
        return 'AP'
    if v < 9.:
        return 'NT'
    if v < 10.:
        return 'SB'
    return 'M'

try:
    from cv_cfg import USERNAME, PASSWORD 
except:
    print('''WARNING: Para recoger calificaciones necesitas un archivo cv_cfg.py con:
USERNAME='Tu nombre de usuario sin @uclm.es'
PASSWORD='Tu contraseña'
''')

if __name__ == '__main__':
    try: os.chdir(os.path.dirname(__file__))
    except: pass
    parser = argparse.ArgumentParser(description='Transfiere los datos de una hoja de CampusVirtual a las actas correspondientes')
    parser.add_argument('--ceil', action='store_true',
                        help='redondea por arriba notas > 4.0')
    parser.add_argument('--noout', action='store_true',
                        help='no produce salidas de ningún tipo')
    parser.add_argument('--fetch', action='store',
                        help='descarga la hoja de Campus Virtual')
    parser.add_argument('--upload', nargs='?', default=None, const='-',
                        help='sube las actas a ActasCalificacion')
    parser.add_argument('--cv', action='store', default='cv.xlsx',
                        help='ruta de la hoja de Campus Virtual')
    parser.add_argument('--actas', nargs='+', required=True,
                        help='renderiza los datos en archivo HTML')
    parser.add_argument('--out-prefix', action='store', default='',
                        help='prefijo de las hojas de salida')
    parser.add_argument('--require', nargs='*',
                        help='tareas requeridas para considerar presentado')
    parser.add_argument('--scale', action='store', type=float, default=1.,
                        help='escalado de las notas (si en CV no son sobre 10)')
    args = parser.parse_args()
    

    if args.fetch or args.upload:
        print('La descarga/subida de archivos con Selenium es lenta, ten paciencia')
        driver = webdriver.Chrome('g:/GitHub/uclm-eii/py/chromedriver.exe')
        actas = ActasCalificacion(USERNAME, PASSWORD, driver)
        actas.authenticate()
    
    if args.fetch:
        cv = CampusVirtual(USERNAME, PASSWORD, driver)
        cv.authenticate()

        print('Descargando actas vacías de', args.fetch)
        for i, f in enumerate(args.actas):
            with open(f, 'wb') as xls:
                print('Descargando acta', args.fetch, i)
                xls.write(actas.download_acta(args.fetch, i))
        print('Descargando calificaciones de', args.fetch)
        with open(args.cv, 'wb') as f:
            f.write(cv.calificaciones(args.fetch))

    if args.ceil:
        grade_ceil = True

    if not args.noout:
        print('Calculando calificaciones de', args.cv)
        calificar_actas(notas = args.cv, 
                        actas = args.actas, 
                        prefijo = args.out_prefix, 
                        requisito = args.require,
                        scale = args.scale)

    if args.upload:
        course = args.upload if args.upload != '-' else args.fetch
        if not course:
            exit()
        print("Subiendo actas de", course)
        for i, fname in enumerate(args.actas):
            print("Subiendo", args.out_prefix + fname)
            actas.upload_acta(course, i, args.out_prefix + fname)
